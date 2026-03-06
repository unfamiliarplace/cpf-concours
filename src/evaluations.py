from __future__ import annotations
from typing import Iterable

import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles.alignment import Alignment
from pathlib import Path
from concours import *

PATH_BASE = Path('./src')
PATH_OUTPUT = PATH_BASE / 'output'
PATH_TEMPLATES = PATH_BASE / 'templates'

PATH_REPORT_TEMPLATE = PATH_TEMPLATES / 'report.xlsx'

PERIOD_ANY = ''

# Utilities

def assign_named_cells(ws, cols: str, row: int, values: Iterable, alignment: str=''):
    """
    >>> assign_named_cells(ws, 'CDEF', 1, (tuple, of, 4, values))
    """
    for (c, v) in zip(cols, values):
        ws[f'{c}{row}'] = v

        if alignment:
            ws[f'{c}{row}'].alignment = Alignment(alignment)

def assign_n_cells(ws, start: str, row: int, values: Iterable, alignment: str=''):
    """
    >>> assign_named_cells(ws, 'C', 1, (tuple, of, values))

    # TODO Simple: does not handle post-Z. Do not supply too many values :)
    """

    for (i, v) in enumerate(values):
        c = chr(ord(start) + i)
        ws[f'{c}{row}'] = v

        if alignment:
            ws[f'{c}{row}'].alignment = Alignment(alignment)

class ConcoursReport:
    c: Concours

    judge_to_sp: dict[Judge, Scorepad]
    judge_to_sp_adj: dict[Judge, Scorepad]
    contestant_to_sp: dict[Contestant, Scorepad]
    contestant_to_sp_adj: dict[Contestant, Scorepad]
    sformat_to_sp: dict[str, Scorepad]
    grade_to_sp: dict[str, Scorepad]
    level_to_sp: dict[str, Scorepad]
    duration_to_sp: dict[int, Scorepad] # Bucket by # of minutes
    duration_to_sp_adjust: dict[int, Scorepad]
    category_to_sp: dict[int, Scorepad]
    category_to_sp_var: dict[int, Scorepad]
    category_to_places: dict[Category, list[Scorepad]]
    school_to_sp_given: dict[School, Scorepad]
    school_to_sp_received: dict[School, Scorepad]

    def __init__(self: ConcoursReport, c: Concours):
        self.c = c

        self.judge_to_sp = {}
        self.judge_to_sp_adj = {}
        self.contestant_to_sp = {}
        self.contestant_to_sp_adj = {}
        self.sformat_to_sp = {}
        self.grade_to_sp = {}
        self.level_to_sp = {}
        self.duration_to_sp = {}
        self.duration_to_sp_adj = {}
        self.category_to_sp = {}
        self.category_to_sp_var = {}
        self.category_to_places = {}
        self.school_to_sp_given = {}
        self.school_to_sp_received = {}

        self.create_scorepads()
    
    def create_scorepads(self: ConcoursReport):

        es = set(filter(lambda e: e.scores != None, self.c.scoreboard.evaluations))
       
        # Must be done first for adjustment
        for cat in self.c.categories:
            self.category_to_sp[cat] = Scorepad(cat, set(filter(lambda e: e.category == cat, es)))            

        # Must precede category places
        for cont in self.c.contestants:
            sp = Scorepad(cont, set(filter(lambda e: e.speech.contestant == cont, es)))
            self.contestant_to_sp[cont] = sp
            self.contestant_to_sp_adj[cont] = sp.adjust_to_category(self.category_to_sp)

        for judge in self.c.judges:
            # Hackish solution to different periods
            judge = Judge(judge.name, judge.school, PERIOD_ANY)

            sp = Scorepad(judge, set(filter(lambda e: e.judge == judge, es)))
            self.judge_to_sp[judge] = sp
            self.judge_to_sp_adj[judge] = sp.adjust_to_category(self.category_to_sp)

        # Places
        for cat in self.c.categories:
            conts = set(filter(lambda sp: sp.item.category == cat, self.contestant_to_sp.values()))
            self.category_to_places[cat] = sorted(conts, key=lambda cont: cont.average(), reverse=True)

        for sformat in SFORMATS:
            self.sformat_to_sp[sformat] = Scorepad(sformat, set(filter(lambda e: e.sformat == sformat, es)))

        for grade in GRADES:
            self.grade_to_sp[grade] = Scorepad(grade, set(filter(lambda e: e.grade == grade, es)))

        for level in LEVELS:
            self.level_to_sp[level] = Scorepad(level, set(filter(lambda e: e.level == level, es)))

        for e in es:
            bucket = round(e.speech.duration / 60)
            self.duration_to_sp.setdefault(bucket, Scorepad(bucket, set())).add_evaluation(e)
        
        for (bucket, sp) in self.duration_to_sp.items():
            self.duration_to_sp_adj[bucket] = sp.adjust_to_category(self.category_to_sp)

        for school in self.c.schools:
            self.school_to_sp_given[school] = Scorepad(school, set(filter(lambda e: e.judge.school == school, es)))
            self.school_to_sp_received[school] = Scorepad(school, set(filter(lambda e: e.contestant.school == school, es)))

    def save(self: ConcoursReport):
        wb = openpyxl.load_workbook(PATH_REPORT_TEMPLATE)

        generics = (
            ('contestants', self.contestant_to_sp),
            ('contestants_adjust', self.contestant_to_sp_adj),
            ('judges', self.judge_to_sp),
            ('judges_adjust', self.judge_to_sp_adj),
            ('categories', self.category_to_sp),
            ('category_variances', self.category_to_sp, 'variance'),
            ('formats', self.sformat_to_sp),
            ('grades', self.grade_to_sp),
            ('levels', self.level_to_sp),
            ('durations', self.duration_to_sp),
            ('durations_adjust', self.duration_to_sp_adj),
            ('schools_given', self.school_to_sp_given),
            ('schools_received', self.school_to_sp_received),
        )

        for args in generics:
            self._save_generic(wb, *args)

        self._save_places(wb)

        path = PATH_OUTPUT / f'statistics_{self.c.name}.xlsx'
        wb.save(path)
    
    def _save_generic(self: ConcoursReport, wb: Workbook, key: str, d: dict[object, set[Scorepad]], mode: str='average'):
        """
        mode can be 'average' or 'variance'
        """

        if mode == 'average':
            cb_total = Scorepad.average
            cb_indiv = Scorepad.averages
        elif mode == 'variance':
            cb_total = Scorepad.variance
            cb_indiv = Scorepad.variances
        else:
            cb_total = lambda _: ''
            cb_indiv = lambda _: [''] * 5

        ws = wb[key]
        for (i, sp) in enumerate(self.sorted_sps_from_dict(d, cb_total)):
            r = i + 2

            spt = sp.filter_traditional()
            spi = sp.filter_impromptu()
            
            ws[f'A{r}'] = str(sp.item)
            ws[f'B{r}'] = sp.n
            ws[f'C{r}'] = cb_total(sp)

            ws[f'D{r}'] = spt.n
            ws[f'E{r}'] = cb_total(spt)
            assign_named_cells(ws, 'FGHIJ', r, cb_indiv(spt))

            ws[f'K{r}'] = spi.n
            ws[f'L{r}'] = cb_total(spi)
            assign_named_cells(ws, 'MNOPQ', r, cb_indiv(spi))

    def _save_places(self: ConcoursReport, wb: Workbook):
        ws = wb['places']
        for (i, cat) in enumerate(sorted(self.category_to_places)):
            r_places = 2 + (2 * i)
            r_scores = r_places + 1
            sps = self.category_to_places[cat]

            # Alternating rows: category / places on one row, scores on next
            ws[f'A{r_places}'] = str(cat)
            assign_n_cells(ws, 'B', r_places, (str(sp.item) for sp in sps))
            assign_n_cells(ws, 'B', r_scores, (sp.average() for sp in sps), alignment='left')

    @staticmethod
    def sorted_sps_from_dict(d: dict[object, Scorepad], cb: callable=None) -> list[Scorepad]:
        if not cb:
            cb = Scorepad.average

        return ConcoursReport.sort_sps(d.values(), cb)

    @staticmethod
    def sort_sps(sps: Iterable[Scorepad], cb: callable=None) -> list[Scorepad]:
        if not cb:
            cb = Scorepad.average

        return sorted(sps, key=cb, reverse=True)

class Scorepad:
    item: object
    evaluations: set[Evaluation]
    n: int

    def __init__(self: Scorepad, item: object, evaluations: set[Evaluation]):
        self.item, self.evaluations = item, evaluations
        self.n = len(self.evaluations)

    def add_evaluation(self: Scorepad, e: Evaluation):
        self.evaluations.add(e)
        self.n = len(self.evaluations)

    def filter_evaluations(self: Scorepad, cb: callable) -> Scorepad:
        return Scorepad(self.item, set(filter(cb, self.evaluations)))

    def filter_traditional(self: Scorepad) -> Scorepad:
        return self.filter_evaluations(lambda e: e.sformat == 'Traditionnel')

    def filter_impromptu(self: Scorepad) -> Scorepad:
        return self.filter_evaluations(lambda e: e.sformat == 'Impromptu')
    
    def adjust_to_category(self: Scorepad, cat_sps: dict[Category, set[Scorepad]]) -> Scorepad:
        """
        Take this Scorepad's evaluations and adjust them by subtracting the average
        of any other Scorepads whose item matches this one's category.
        """

        new_es = []
        for e in self.evaluations:
            cat_avgs = cat_sps[e.category].averages()

            new_scores = []
            for i in range(5):
                new_scores.append(e.scores[i] - cat_avgs[i])
            
            new_e = Evaluation(e.judge, e.speech, new_scores)
            new_es.append(new_e)
        
        return Scorepad(self.item, new_es)
    
    def average(self: Scorepad) -> float:
        if not self.n:
            return 0.0
        else:
            return round(sum(sum(e.scores) for e in self.evaluations) / self.n, 1)
    
    def averages(self: Scorepad) -> list[float]:
        totals = [0.0, 0.0, 0.0, 0.0, 0.0]
        if not self.n:
            return totals

        for e in self.evaluations:
            for i in range(5):
                totals[i] += e.scores[i]

        for i in range(5):
                totals[i] = round(totals[i] / self.n, 1)

        return totals
        
    def variance(self: Scorepad) -> float:
        if not self.n:
            return 0.0
        else:
            total = 0.0

            mean = self.average()
            for e in self.evaluations:
                diff = sum(e.scores) - mean
                total += diff ** 2
        
            return round(total / self.n, 1)
        
    def variances(self: Scorepad) -> list[float]:
        totals = [0.0, 0.0, 0.0, 0.0, 0.0]
        if not self.n:
            return totals
        
        means = self.averages()

        for e in self.evaluations:
            for i in range(5):
                diff = e.scores[i] - means[i]
                totals[i] += diff ** 2

        for i in range(5):
                totals[i] = round(totals[i] / self.n, 1)

        return totals
